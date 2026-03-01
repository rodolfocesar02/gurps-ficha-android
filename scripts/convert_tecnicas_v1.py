#!/usr/bin/env python3
import argparse
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    lowered = ascii_only.lower()
    lowered = re.sub(r"[^a-z0-9]+", "_", lowered)
    lowered = re.sub(r"_+", "_", lowered).strip("_")
    return lowered or "tecnica"


def parse_page(raw: str):
    if not raw:
        return None
    numbers = re.findall(r"\d+", raw)
    if not numbers:
        return None
    return int(numbers[0])


def header_indexes(headers):
    norm = {h.lower(): idx for idx, h in enumerate(headers)}

    def pick(*aliases):
        for alias in aliases:
            if alias in norm:
                return norm[alias]
        return None

    pagina_idx = pick("página", "pagina")
    nome_idx = pick("nome")
    dificuldade_idx = pick("dificuldade")
    pre_definido_idx = pick("pré-definido", "pre-definido", "pre definido")
    pre_requisito_idx = pick("pré-requisito", "pre-requisito", "pre requisito")
    descricao_idx = pick("descrição", "descricao")
    required_map = {
        "página": pagina_idx,
        "nome": nome_idx,
        "dificuldade": dificuldade_idx,
        "pré-definido": pre_definido_idx,
        "pré-requisito": pre_requisito_idx,
        "descrição": descricao_idx,
    }
    missing = [field for field, idx in required_map.items() if idx is None]
    if missing:
        raise SystemExit(f"Cabeçalhos obrigatórios ausentes: {missing}. Cabeçalhos: {headers}")

    optional_mod = norm.get("modificadores")
    return {
        "pagina": pagina_idx,
        "nome": nome_idx,
        "dificuldade": dificuldade_idx,
        "preDefinido": pre_definido_idx,
        "preRequisito": pre_requisito_idx,
        "descricao": descricao_idx,
        "modificadores": optional_mod,
    }


def convert(inputs: list[Path], sources: list[str], output_json: Path):
    if len(inputs) != len(sources):
        raise SystemExit("--input e --source devem ter a mesma quantidade de ocorrências.")

    items = []
    used_ids = set()
    for input_xlsx, source_book in zip(inputs, sources):
        wb = load_workbook(input_xlsx, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [as_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        idx = header_indexes(headers)

        source_slug = slugify(source_book)
        for r in range(2, ws.max_row + 1):
            values = [as_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            nome = values[idx["nome"]]
            if not nome:
                continue

            page_raw = values[idx["pagina"]]
            dificuldade = values[idx["dificuldade"]]
            pre_definido = values[idx["preDefinido"]]
            pre_requisito = values[idx["preRequisito"]]
            descricao = values[idx["descricao"]]
            modificadores = ""
            if idx["modificadores"] is not None:
                modificadores = values[idx["modificadores"]]

            base = f"{source_slug}_{slugify(nome)}"
            item_id = base
            suffix = 2
            while item_id in used_ids:
                item_id = f"{base}_{suffix}"
                suffix += 1
            used_ids.add(item_id)

            items.append(
                {
                    "id": item_id,
                    "nome": nome,
                    "pagina": parse_page(page_raw),
                    "paginaRaw": page_raw,
                    "dificuldadeRaw": dificuldade,
                    "preDefinidoRaw": pre_definido,
                    "preRequisitoRaw": pre_requisito,
                    "descricao": descricao,
                    "modificadores": modificadores,
                    "sourceBook": source_book,
                    "sourceFile": input_xlsx.name,
                }
            )

    payload = {
        "version": 1,
        "kind": "tecnicas_catalogo_v1",
        "sourceFiles": [str(p) for p in inputs],
        "generatedAtUtc": datetime.now(timezone.utc).isoformat(),
        "totalItems": len(items),
        "items": items,
    }

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"OK: {len(items)} itens => {output_json}")


def main():
    parser = argparse.ArgumentParser(description="Converte técnicas (múltiplas planilhas) para JSON v1")
    parser.add_argument("--input", action="append", required=True, help="Caminho do XLSX (repita para múltiplos)")
    parser.add_argument("--source", action="append", required=True, help="Nome do livro/origem (na mesma ordem)")
    parser.add_argument("--output", required=True, help="Caminho do JSON de saída")
    args = parser.parse_args()
    convert([Path(p) for p in args.input], args.source, Path(args.output))


if __name__ == "__main__":
    main()
