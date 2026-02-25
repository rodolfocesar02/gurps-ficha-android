#!/usr/bin/env python3
import argparse
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_HEADERS = [
    "tipo",
    "categoria",
    "grupo",
    "nt",
    "nome",
    "tipo_dano",
    "precisao",
    "alcance_distancia",
    "peso",
    "Cdt",
    "tiros",
    "custo",
    "st_minimo",
    "Magnitude",
    "Recuo",
    "Cl",
    "observacoes",
]


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    lowered = ascii_only.lower()
    lowered = re.sub(r"[^a-z0-9]+", "_", lowered)
    lowered = re.sub(r"_+", "_", lowered).strip("_")
    return lowered or "arma_distancia"


def as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def convert(input_xlsx: Path, output_json: Path):
    wb = load_workbook(input_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [as_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    if headers != EXPECTED_HEADERS:
        raise SystemExit(
            "Cabeçalhos inesperados na planilha.\n"
            f"Esperado: {EXPECTED_HEADERS}\n"
            f"Obtido:   {headers}"
        )

    items = []
    used_ids = set()

    for r in range(2, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        row = dict(zip(headers, values))
        nome = as_text(row["nome"])
        if not nome:
            continue

        base_id = slugify(nome)
        final_id = base_id
        suffix = 2
        while final_id in used_ids:
            final_id = f"{base_id}_{suffix}"
            suffix += 1
        used_ids.add(final_id)

        item = {
            "id": final_id,
            "tipo": as_text(row["tipo"]) or "arma_distancia",
            "categoria": as_text(row["categoria"]),
            "grupo": as_text(row["grupo"]),
            "nt": as_text(row["nt"]),
            "nome": nome,
            "tipo_dano": as_text(row["tipo_dano"]),
            "precisao": as_text(row["precisao"]),
            "alcance_distancia": as_text(row["alcance_distancia"]),
            "peso": as_text(row["peso"]),
            "Cdt": as_text(row["Cdt"]),
            "tiros": as_text(row["tiros"]),
            "custo": as_text(row["custo"]),
            "st_minimo": as_text(row["st_minimo"]),
            "Magnitude": as_text(row["Magnitude"]),
            "Recuo": as_text(row["Recuo"]),
            "Cl": as_text(row["Cl"]),
            "observacoes": as_text(row["observacoes"]),
            "source": "xlsx",
            "reviewFlags": [],
            "rowNumber": r,
        }
        items.append(item)

    payload = {
        "version": 1,
        "kind": "armas_distancia_raw_from_xlsx",
        "sourceFile": str(input_xlsx),
        "generatedAtUtc": datetime.now(timezone.utc).isoformat(),
        "totalItems": len(items),
        "items": items,
    }

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"OK: {len(items)} itens => {output_json}")


def main():
    parser = argparse.ArgumentParser(description="Converte planilha de armas à distância para JSON raw v1")
    parser.add_argument("--input", required=True, help="Caminho do .xlsx")
    parser.add_argument("--output", required=True, help="Caminho do JSON de saída")
    args = parser.parse_args()
    convert(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
