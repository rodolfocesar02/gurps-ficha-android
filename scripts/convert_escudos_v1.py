#!/usr/bin/env python3
import argparse
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_HEADERS = ["NT", "Escudo", "BD", "Custo", "Peso", "RD/PV", "CL", "Observações"]


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    lowered = ascii_only.lower()
    lowered = re.sub(r"[^a-z0-9]+", "_", lowered)
    lowered = re.sub(r"_+", "_", lowered).strip("_")
    return lowered or "escudo"


def as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_int(value: str):
    if not value:
        return None
    return int(value) if re.fullmatch(r"-?\d+", value) else None


def parse_float_pt(value: str):
    if not value:
        return None
    cleaned = value.replace(".", "").replace(",", ".")
    return float(cleaned) if re.fullmatch(r"-?\d+(\.\d+)?", cleaned) else None


def parse_money(value: str):
    if not value:
        return None
    cleaned = value.replace("$", "").replace(".", "").replace(",", ".").strip()
    return float(cleaned) if re.fullmatch(r"-?\d+(\.\d+)?", cleaned) else None


def parse_st(value: str):
    raw = as_text(value)
    if not raw:
        return None
    cleaned = raw.replace("†", "").replace("‡", "").strip()
    return parse_int(cleaned)


def convert(input_xlsx: Path, output_json: Path):
    wb = load_workbook(input_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [as_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise SystemExit(f"Cabeçalhos inesperados. Esperado: {EXPECTED_HEADERS} | Obtido: {headers}")

    items = []
    used_ids = set()
    for r in range(2, ws.max_row + 1):
        nt_raw = as_text(ws.cell(r, 1).value)
        nome = as_text(ws.cell(r, 2).value)
        db_raw = as_text(ws.cell(r, 3).value)
        custo_raw = as_text(ws.cell(r, 4).value)
        peso_raw = as_text(ws.cell(r, 5).value)
        rd_pv_raw = as_text(ws.cell(r, 6).value)
        cl_raw = as_text(ws.cell(r, 7).value)
        obs_raw = as_text(ws.cell(r, 8).value)

        if not nome:
            continue

        base_id = slugify(nome)
        final_id = base_id
        suffix = 2
        while final_id in used_ids:
            final_id = f"{base_id}_{suffix}"
            suffix += 1
        used_ids.add(final_id)

        rd = None
        pv = None
        if "/" in rd_pv_raw:
            a, b = [x.strip() for x in rd_pv_raw.split("/", 1)]
            rd = parse_int(a)
            pv = parse_int(b)

        item = {
            "id": final_id,
            "nome": nome,
            "ntRaw": nt_raw,
            "nt": parse_int(nt_raw),
            "db": parse_int(db_raw),
            "custoRaw": custo_raw,
            "custo": parse_money(custo_raw),
            "pesoRaw": peso_raw,
            "pesoKg": parse_float_pt(peso_raw),
            "rdPvRaw": rd_pv_raw,
            "rd": rd,
            "pv": pv,
            "clRaw": cl_raw,
            "cl": parse_int(cl_raw),
            "observacoes": obs_raw,
            "tipoEquipamento": "ESCUDO",
            "rowNumber": r,
        }
        items.append(item)

    payload = {
        "version": 1,
        "kind": "escudos_v1",
        "sourceFile": str(input_xlsx),
        "generatedAtUtc": datetime.now(timezone.utc).isoformat(),
        "totalItems": len(items),
        "items": items,
    }
    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"OK: {len(items)} itens => {output_json}")


def main():
    parser = argparse.ArgumentParser(description="Converte planilha de escudos para escudos.v1.json")
    parser.add_argument("--input", required=True, help="Caminho da planilha de escudos")
    parser.add_argument("--output", required=True, help="Caminho do JSON de saída")
    args = parser.parse_args()
    convert(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
