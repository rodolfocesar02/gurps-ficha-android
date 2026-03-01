#!/usr/bin/env python3
import argparse
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook


MOJIBAKE_TOKENS = ["Ã", "Â", "�", "â€", "â€œ", "â€", "â€¢"]


@dataclass
class CatalogIndex:
    source: str
    names_norm: List[Tuple[str, str]]


def as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def norm(value: str) -> str:
    text = unicodedata.normalize("NFD", value or "")
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.lower().replace("-", " ")
    text = re.sub(r"[^a-z0-9\s/+()]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def slugify(value: str) -> str:
    t = norm(value)
    t = re.sub(r"\(\s*[†*]?\s*\)", "", t)
    t = t.replace("/", " ")
    t = re.sub(r"[^a-z0-9]+", "_", t)
    t = re.sub(r"_+", "_", t).strip("_")
    return t or "pericia"


def remove_marcadores_nome(nome: str) -> str:
    t = nome.replace("†", "")
    t = re.sub(r"\(\s*[†*]\s*\)", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def parse_tipo(raw_tipo: str) -> Dict:
    tipo = as_text(raw_tipo)
    if not tipo:
        return {
            "atributoBase": None,
            "dificuldadeFixa": None,
            "dificuldadeVariavel": False,
            "tipoValido": False,
            "motivo": "tipo_vazio",
        }

    parts = [p.strip() for p in tipo.split("/", 1)]
    if len(parts) != 2:
        return {
            "atributoBase": None,
            "dificuldadeFixa": None,
            "dificuldadeVariavel": False,
            "tipoValido": False,
            "motivo": "formato_sem_barra",
        }

    attr_raw, diff_raw = parts[0], parts[1]
    attr_n = norm(attr_raw)
    diff_n = norm(diff_raw)

    attr_map = {
        "st": "ST",
        "dx": "DX",
        "iq": "IQ",
        "ht": "HT",
        "per": "PER",
        "vontade": "VON",
        "von": "VON",
    }
    diff_map = {
        "facil": "F",
        "media": "M",
        "dificil": "D",
        "muito dificil": "MD",
        "variavel": None,
    }

    atributo = attr_map.get(attr_n)
    if atributo is None:
        return {
            "atributoBase": None,
            "dificuldadeFixa": None,
            "dificuldadeVariavel": False,
            "tipoValido": False,
            "motivo": f"atributo_invalido:{attr_raw}",
        }

    if diff_n not in diff_map:
        return {
            "atributoBase": atributo,
            "dificuldadeFixa": None,
            "dificuldadeVariavel": False,
            "tipoValido": False,
            "motivo": f"dificuldade_invalida:{diff_raw}",
        }

    if diff_n == "variavel":
        return {
            "atributoBase": atributo,
            "dificuldadeFixa": None,
            "dificuldadeVariavel": True,
            "tipoValido": True,
            "motivo": "ok_variavel",
        }

    return {
        "atributoBase": atributo,
        "dificuldadeFixa": diff_map[diff_n],
        "dificuldadeVariavel": False,
        "tipoValido": True,
        "motivo": "ok",
    }


def classificar_especializacao(nome_raw: str, predef_raw: str, prereq_raw: str, mod_raw: str) -> Dict:
    nome = as_text(nome_raw)
    nome_n = norm(nome)
    texto = " | ".join([as_text(predef_raw), as_text(prereq_raw), as_text(mod_raw)])
    texto_n = norm(texto)

    motivos = []
    exige = False

    if "†" in nome or re.search(r"\(\s*†\s*\)", nome):
        exige = True
        motivos.append("marcador_†")

    if "(várias" in nome.lower() or "(varias" in nome.lower():
        exige = True
        motivos.append("nome_varias")

    if "entre especializ" in texto_n or "outra especializacao" in texto_n:
        exige = True
        motivos.append("texto_entre_especializacoes")

    familias_genericas = [
        "assuntos atuais/nt",
        "conducao/nt",
        "pilotagem/nt",
        "sacar rapido",
        "armas de fogo/nt",
        "armas de feixe/nt",
        "canhoneiro/nt",
        "conhecimento oculto",
        "sobrevivencia",
        "trato social",
        "ritual religioso",
        "arma de arremesso",
        "ataque inato",
        "pericia profissional",
        "pericias de passatempo",
        "arte ou esporte de combate",
        "artista",
    ]
    if any(norm(f) in nome_n for f in familias_genericas):
        exige = True
        motivos.append("familia_generica")

    if not motivos:
        motivos.append("sem_indicio_explicito")

    return {"exigeEspecializacao": exige, "motivos": motivos}


def load_json_any(path: Path):
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict) and isinstance(payload.get("items"), list):
        return payload["items"]
    return []


def index_catalog(names_source: str, items: List[Dict], name_field: str = "nome") -> CatalogIndex:
    pairs = []
    for it in items:
        nome = as_text(it.get(name_field, ""))
        if not nome:
            continue
        nn = norm(nome)
        if nn:
            pairs.append((nome, nn))
    pairs.sort(key=lambda x: len(x[1]), reverse=True)
    return CatalogIndex(source=names_source, names_norm=pairs)


def find_links(text: str, catalog: CatalogIndex, limit: int = 8) -> List[str]:
    t = f" {norm(text)} "
    found = []
    seen = set()
    for nome, nn in catalog.names_norm:
        if len(nn) < 4:
            continue
        token = f" {nn} "
        if token in t and nome not in seen:
            found.append(nome)
            seen.add(nome)
            if len(found) >= limit:
                break
    return found


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera preview tecnico da Tabbela de pericias_V2.xlsx e auditoria de ligacoes.")
    parser.add_argument("--input-xlsx", required=True)
    parser.add_argument("--assets-dir", default="app/src/main/assets")
    parser.add_argument("--preview-out", default="scripts/reports/pericias_v2_preview.json")
    parser.add_argument("--quality-out", default="scripts/reports/pericias_v2_quality_report.json")
    args = parser.parse_args()

    xlsx = Path(args.input_xlsx)
    assets = Path(args.assets_dir)

    wb = load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [as_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    vantagens = load_json_any(assets / "vantagens.v3.json")
    desvantagens = load_json_any(assets / "desvantagens.v2.json")
    magias = load_json_any(assets / "magias.json")
    tecnicas = load_json_any(assets / "tecnicas.v1.json")

    idx_v = index_catalog("vantagens", vantagens)
    idx_d = index_catalog("desvantagens", desvantagens)
    idx_m = index_catalog("magias", magias)
    idx_t = index_catalog("tecnicas", tecnicas)

    preview_items = []
    stats = {
        "totalRows": 0,
        "tipoValido": 0,
        "tipoInvalido": 0,
        "exigeEspecializacaoTrue": 0,
        "exigeEspecializacaoFalse": 0,
        "mojibakeSuspeitas": 0,
        "preRequisitoVazio": 0,
        "preDefinidoVazio": 0,
        "descricaoVazia": 0,
        "modificadoresVazio": 0,
        "linksEstritos": {"vantagens": 0, "desvantagens": 0, "magias": 0, "tecnicas": 0},
        "linksAmplos": {"vantagens": 0, "desvantagens": 0, "magias": 0, "tecnicas": 0},
    }

    tipo_invalidos = []
    ligacoes_amostras = []

    for r in range(2, ws.max_row + 1):
        pagina = as_text(ws.cell(r, 1).value)
        nome = as_text(ws.cell(r, 2).value)
        tipo_raw = as_text(ws.cell(r, 3).value)
        predef_raw = as_text(ws.cell(r, 4).value)
        prereq_raw = as_text(ws.cell(r, 5).value)
        desc = as_text(ws.cell(r, 6).value)
        mod = as_text(ws.cell(r, 7).value)

        if not any([pagina, nome, tipo_raw, predef_raw, prereq_raw, desc, mod]):
            continue

        stats["totalRows"] += 1

        tipo = parse_tipo(tipo_raw)
        if tipo["tipoValido"]:
            stats["tipoValido"] += 1
        else:
            stats["tipoInvalido"] += 1
            tipo_invalidos.append({"row": r, "nome": nome, "tipoRaw": tipo_raw, "motivo": tipo["motivo"]})

        esp = classificar_especializacao(nome, predef_raw, prereq_raw, mod)
        if esp["exigeEspecializacao"]:
            stats["exigeEspecializacaoTrue"] += 1
        else:
            stats["exigeEspecializacaoFalse"] += 1

        blob_estrito = " | ".join([predef_raw, prereq_raw])
        blob_amplo = " | ".join([predef_raw, prereq_raw, desc, mod])

        moj = {t: blob_amplo.count(t) for t in MOJIBAKE_TOKENS if blob_amplo.count(t) > 0}
        if moj:
            stats["mojibakeSuspeitas"] += 1

        if predef_raw in {"", "-", "—"}:
            stats["preDefinidoVazio"] += 1
        if prereq_raw in {"", "-", "—"}:
            stats["preRequisitoVazio"] += 1
        if desc in {"", "-", "—"}:
            stats["descricaoVazia"] += 1
        if mod in {"", "-", "—"}:
            stats["modificadoresVazio"] += 1

        links_v_estrito = find_links(blob_estrito, idx_v)
        links_d_estrito = find_links(blob_estrito, idx_d)
        links_m_estrito = find_links(blob_estrito, idx_m)
        links_t_estrito = find_links(blob_estrito, idx_t)

        links_v_amplo = find_links(blob_amplo, idx_v)
        links_d_amplo = find_links(blob_amplo, idx_d)
        links_m_amplo = find_links(blob_amplo, idx_m)
        links_t_amplo = find_links(blob_amplo, idx_t)

        if links_v_estrito:
            stats["linksEstritos"]["vantagens"] += 1
        if links_d_estrito:
            stats["linksEstritos"]["desvantagens"] += 1
        if links_m_estrito:
            stats["linksEstritos"]["magias"] += 1
        if links_t_estrito:
            stats["linksEstritos"]["tecnicas"] += 1

        if links_v_amplo:
            stats["linksAmplos"]["vantagens"] += 1
        if links_d_amplo:
            stats["linksAmplos"]["desvantagens"] += 1
        if links_m_amplo:
            stats["linksAmplos"]["magias"] += 1
        if links_t_amplo:
            stats["linksAmplos"]["tecnicas"] += 1

        if any([links_v_estrito, links_d_estrito, links_m_estrito, links_t_estrito]) and len(ligacoes_amostras) < 30:
            ligacoes_amostras.append(
                {
                    "row": r,
                    "nome": nome,
                    "preRequisitoRaw": prereq_raw,
                    "preDefinidoRaw": predef_raw,
                    "linksEstritos": {
                        "vantagens": links_v_estrito,
                        "desvantagens": links_d_estrito,
                        "magias": links_m_estrito,
                        "tecnicas": links_t_estrito,
                    },
                }
            )

        preview_items.append(
            {
                "row": r,
                "idPreview": slugify(remove_marcadores_nome(nome)),
                "paginaRaw": pagina,
                "nomeOriginal": nome,
                "nomeCanonico": remove_marcadores_nome(nome),
                "tipoRaw": tipo_raw,
                "parseTipo": tipo,
                "classificacaoEspecializacao": esp,
                "preDefinidoRaw": predef_raw,
                "preRequisitoRaw": prereq_raw,
                "descricao": desc,
                "modificadores": mod,
                "linksCruzados": {
                    "estrito": {
                        "vantagens": links_v_estrito,
                        "desvantagens": links_d_estrito,
                        "magias": links_m_estrito,
                        "tecnicas": links_t_estrito,
                    },
                    "amplo": {
                        "vantagens": links_v_amplo,
                        "desvantagens": links_d_amplo,
                        "magias": links_m_amplo,
                        "tecnicas": links_t_amplo,
                    },
                },
                "sinaisMojibake": moj,
            }
        )

    quality = {
        "input": str(xlsx),
        "headers": headers,
        "summary": stats,
        "tipoInvalidos": tipo_invalidos[:120],
        "amostrasLigacoesEstritas": ligacoes_amostras,
        "recomendacoes": [
            "Normalizar Tipo: mapear Vontade->VON, Per->PER e Variavel para dificuldadeVariavel=true.",
            "Transformar familias com especializacao (marcador † ou 'varias') em exigeEspecializacao=true com opcao de especializacao no app.",
            "Estruturar preDefinidoRaw em regras parseaveis (atributo/pericia/modificador) para calculo automatico.",
            "Estruturar preRequisitoRaw em termos parseaveis (OR/AND; vantagens; niveis minimos) para validacao automatica robusta.",
            "Padronizar nomes para reduzir divergencia com pericias.json atual e melhorar casamento tecnico/pericia base.",
        ],
    }

    preview_payload = {
        "version": 2,
        "kind": "pericias_v2_preview",
        "sourceFile": str(xlsx),
        "totalItems": len(preview_items),
        "items": preview_items,
    }

    preview_out = Path(args.preview_out)
    quality_out = Path(args.quality_out)
    preview_out.parent.mkdir(parents=True, exist_ok=True)
    quality_out.parent.mkdir(parents=True, exist_ok=True)

    preview_out.write_text(json.dumps(preview_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    quality_out.write_text(json.dumps(quality, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print("=== Preview Tecnico Pericias V2 ===")
    print(f"Input: {xlsx}")
    print(f"Items: {len(preview_items)}")
    print(f"Tipo valido: {stats['tipoValido']} | Tipo invalido: {stats['tipoInvalido']}")
    print(
        "Links estritos (preDef+preReq) -> "
        f"vantagens:{stats['linksEstritos']['vantagens']} "
        f"desvantagens:{stats['linksEstritos']['desvantagens']} "
        f"magias:{stats['linksEstritos']['magias']} "
        f"tecnicas:{stats['linksEstritos']['tecnicas']}"
    )
    print(
        "Links amplos (inclui desc+mods) -> "
        f"vantagens:{stats['linksAmplos']['vantagens']} "
        f"desvantagens:{stats['linksAmplos']['desvantagens']} "
        f"magias:{stats['linksAmplos']['magias']} "
        f"tecnicas:{stats['linksAmplos']['tecnicas']}"
    )
    print(f"Preview: {preview_out}")
    print(f"Quality: {quality_out}")


if __name__ == "__main__":
    main()
