#!/usr/bin/env python3
import argparse
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    lowered = ascii_only.lower()
    lowered = re.sub(r"[^a-z0-9]+", "_", lowered)
    lowered = re.sub(r"_+", "_", lowered).strip("_")
    return lowered or "arma"


def as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
    return str(value).strip()


def parse_int_like(value: str):
    if not value:
        return None
    m = re.fullmatch(r"-?\d+", value)
    if not m:
        return None
    return int(value)


def parse_decimal_ptbr(value: str):
    if not value:
        return None
    cleaned = value.strip().replace(".", "").replace(",", ".")
    m = re.fullmatch(r"-?\d+(\.\d+)?", cleaned)
    if not m:
        return None
    return float(cleaned)


def split_modes(raw: str):
    if not raw:
        return []
    return [part.strip() for part in raw.split("/") if part.strip()]


def parse_cost_token(token: str):
    raw = token.strip()
    if not raw:
        return {"raw": raw, "kind": "empty", "value": None}

    lower = raw.lower()
    if lower in {"—", "-", "var.", "var"}:
        return {"raw": raw, "kind": "special", "value": None}

    sign = -1 if raw.startswith("-") else 1
    has_plus_prefix = raw.startswith("+")
    cleaned = raw.replace("$", "").replace("+", "").replace("-", "").strip()
    numeric = parse_decimal_ptbr(cleaned)
    if numeric is None:
        return {"raw": raw, "kind": "special", "value": None}

    return {
        "raw": raw,
        "kind": "numeric",
        "value": sign * numeric,
        "hasPlusPrefix": has_plus_prefix,
    }


def normalize_row(headers, values, used_ids):
    row = dict(zip(headers, values))

    tipo = as_text(row.get("tipo"))
    categoria = as_text(row.get("categoria"))
    grupo = as_text(row.get("grupo"))
    nt_raw = as_text(row.get("nt"))
    nome = as_text(row.get("nome"))
    dano = as_text(row.get("dano"))
    alcance_corpo = as_text(row.get("alcance_corpo"))
    aparar = as_text(row.get("aparar"))
    custo = as_text(row.get("custo"))
    peso = as_text(row.get("peso"))
    st_minimo = as_text(row.get("st_minimo"))
    observacoes = as_text(row.get("observacoes"))

    if not nome:
        return None

    base_id = slugify(nome)
    final_id = base_id
    suffix = 2
    while final_id in used_ids:
        final_id = f"{base_id}_{suffix}"
        suffix += 1
    used_ids.add(final_id)

    nt = parse_int_like(nt_raw)

    st_clean = st_minimo.replace("†", "").replace("‡", "").strip()
    st_num = parse_int_like(st_clean)
    st_flags = []
    if "†" in st_minimo:
        st_flags.append("dagger")
    if "‡" in st_minimo:
        st_flags.append("double_dagger")

    alcance_modos = split_modes(alcance_corpo)
    aparar_modos = split_modes(aparar)
    custo_modos = [parse_cost_token(x) for x in split_modes(custo)]
    peso_modos_raw = split_modes(peso)
    peso_modos = []
    for token in peso_modos_raw:
        parsed = parse_decimal_ptbr(token)
        peso_modos.append({"raw": token, "kg": parsed})

    return {
        "id": final_id,
        "tipo": tipo,
        "categoria": categoria,
        "grupo": grupo,
        "nome": nome,
        "ntRaw": nt_raw,
        "nt": nt,
        "danoRaw": dano,
        "alcanceCorpoRaw": alcance_corpo,
        "alcanceCorpoModos": alcance_modos,
        "apararRaw": aparar,
        "apararModos": aparar_modos,
        "custoRaw": custo,
        "custoModos": custo_modos,
        "pesoRaw": peso,
        "pesoModos": peso_modos,
        "stMinimoRaw": st_minimo,
        "stMinimo": st_num,
        "stFlags": st_flags,
        "observacoes": observacoes,
    }


def convert(input_xlsx: Path, output_json: Path):
    wb = load_workbook(input_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [as_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    expected = [
        "tipo",
        "categoria",
        "grupo",
        "nt",
        "nome",
        "dano",
        "alcance_corpo",
        "aparar",
        "custo",
        "peso",
        "st_minimo",
        "observacoes",
    ]
    if headers[: len(expected)] != expected:
        raise SystemExit(
            "Cabeçalhos inesperados na planilha.\n"
            f"Esperado: {expected}\n"
            f"Obtido:   {headers}"
        )

    used_ids = set()
    items = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        item = normalize_row(headers, values, used_ids)
        if item is None:
            continue
        item["rowNumber"] = r
        items.append(item)

    output_json.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": 1,
        "kind": "armas_corpo_a_corpo_raw",
        "sourceFile": str(input_xlsx),
        "generatedAtUtc": datetime.now(timezone.utc).isoformat(),
        "totalItems": len(items),
        "items": items,
    }
    output_json.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"OK: {len(items)} itens => {output_json}")


def main():
    parser = argparse.ArgumentParser(
        description="Converte planilha de armas corpo a corpo para JSON raw v1"
    )
    parser.add_argument("--input", required=True, help="Caminho do .xlsx")
    parser.add_argument("--output", required=True, help="Caminho do JSON de saída")
    args = parser.parse_args()
    convert(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
