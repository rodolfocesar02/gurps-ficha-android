#!/usr/bin/env python3
import argparse
import json
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


def as_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def norm(v: str) -> str:
    t = unicodedata.normalize("NFD", v or "")
    t = "".join(ch for ch in t if unicodedata.category(ch) != "Mn")
    t = t.lower().replace("-", " ")
    t = re.sub(r"[^a-z0-9\s/+().]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def slugify(v: str) -> str:
    t = norm(v)
    t = re.sub(r"\(\s*[†*]\s*\)", "", t)
    t = re.sub(r"[^a-z0-9]+", "_", t).strip("_")
    return t or "pericia"


def parse_tipo(raw: str) -> Dict:
    t = as_text(raw)
    parts = [p.strip() for p in t.split("/", 1)]
    attr_map = {
        "st": "ST",
        "dx": "DX",
        "iq": "IQ",
        "ht": "HT",
        "per": "PER",
        "vontade": "VON",
        "von": "VON",
    }
    diff_map = {
        "facil": "F",
        "media": "M",
        "dificil": "D",
        "muito dificil": "MD",
        "variavel": None,
    }

    if len(parts) != 2:
        return {
            "valid": False,
            "attributeMode": "unknown",
            "attributeOptions": [],
            "difficultyMode": "unknown",
            "difficulty": None,
            "raw": t,
            "reason": "sem_barra",
        }

    attr_raw, diff_raw = parts
    attr_n = norm(attr_raw)
    diff_n = norm(diff_raw)

    # casos "DX ou IQ"
    attr_options = []
    if " ou " in attr_n:
        sub = [p.strip() for p in attr_n.split(" ou ")]
        for s in sub:
            mapped = attr_map.get(s)
            if mapped:
                attr_options.append(mapped)
    else:
        mapped = attr_map.get(attr_n)
        if mapped:
            attr_options.append(mapped)

    if not attr_options:
        return {
            "valid": False,
            "attributeMode": "unknown",
            "attributeOptions": [],
            "difficultyMode": "unknown",
            "difficulty": None,
            "raw": t,
            "reason": f"atributo_invalido:{attr_raw}",
        }

    difficulty = diff_map.get(diff_n)
    if diff_n not in diff_map:
        return {
            "valid": False,
            "attributeMode": "choice" if len(attr_options) > 1 else "fixed",
            "attributeOptions": attr_options,
            "difficultyMode": "unknown",
            "difficulty": None,
            "raw": t,
            "reason": f"dificuldade_invalida:{diff_raw}",
        }

    return {
        "valid": True,
        "attributeMode": "choice" if len(attr_options) > 1 else "fixed",
        "attributeOptions": attr_options,
        "difficultyMode": "variable" if difficulty is None else "fixed",
        "difficulty": difficulty,
        "raw": t,
    }


def parse_prerequisitos(raw: str, vantagens_norm: Dict[str, str]) -> Dict:
    txt = as_text(raw)
    if txt in {"", "-", "—"}:
        return {
            "raw": txt,
            "kind": "none",
            "allowWithoutPrerequisite": True,
            "logic": None,
            "tokens": [],
        }

    parts_and = [p.strip() for p in re.split(r";", txt) if p.strip()]
    and_groups = []
    tokens = []

    for part in parts_and:
        ors = [o.strip() for o in re.split(r"\bou\b", part, flags=re.IGNORECASE) if o.strip()]
        or_items = []
        for item in ors:
            inorm = norm(item)
            token = {"raw": item, "normalized": inorm, "type": "text", "value": item}

            # vantagem obrigatoria
            m_vant = re.search(r"vantagem\s+(.+)$", item, flags=re.IGNORECASE)
            if m_vant:
                vname = m_vant.group(1).strip()
                token = {"raw": item, "normalized": inorm, "type": "required_advantage", "value": vname}
            else:
                # nivel minimo de pericia ex.: Arco 18+
                m_lvl = re.search(r"(.+?)\s+(\d+)\+", item)
                if m_lvl:
                    token = {
                        "raw": item,
                        "normalized": inorm,
                        "type": "skill_min_level",
                        "skill": m_lvl.group(1).strip(),
                        "minLevel": int(m_lvl.group(2)),
                    }

            # valida se bate com vantagens catalogadas
            if token.get("type") == "required_advantage":
                vnorm = norm(token["value"])
                if vnorm in vantagens_norm:
                    token["catalogMatch"] = vantagens_norm[vnorm]

            or_items.append(token)
            tokens.append(token)

        and_groups.append({"or": or_items})

    kind = "structured" if and_groups else "text"
    return {
        "raw": txt,
        "kind": kind,
        "allowWithoutPrerequisite": False,
        "logic": {"and": and_groups} if and_groups else None,
        "tokens": tokens,
    }


def parse_predefinido(raw: str) -> Dict:
    txt = as_text(raw)
    if txt in {"", "-", "—"}:
        return {
            "raw": txt,
            "kind": "empty",
            "onZeroPoints": "use_default_predefinido",
            "parsed": [],
        }

    # normaliza separadores e OR
    options = [o.strip() for o in re.split(r"\bou\b", txt, flags=re.IGNORECASE) if o.strip()]
    parsed = []

    for opt in options:
        # ex: DX-6, IQ-5, Acrobacia-4, outra especialização -2
        m = re.match(r"(.+?)\s*([+-]\s*\d+)", opt)
        if m:
            base = m.group(1).strip()
            mod = int(m.group(2).replace(" ", ""))
            base_n = norm(base)
            if base_n in {"st", "dx", "iq", "ht", "per", "vontade", "von"}:
                ptype = "attribute"
            else:
                ptype = "skill_or_rule"
            parsed.append({
                "raw": opt,
                "type": ptype,
                "base": base,
                "modifier": mod,
            })
        else:
            parsed.append({"raw": opt, "type": "text", "base": opt, "modifier": None})

    return {
        "raw": txt,
        "kind": "parsed",
        "onZeroPoints": "use_parsed_predefinido",
        "parsed": parsed,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera pericias_v2_rules_map.json pronto para motor de regras.")
    parser.add_argument("--input-xlsx", required=True)
    parser.add_argument("--assets-dir", default="app/src/main/assets")
    parser.add_argument("--output", default="scripts/reports/pericias_v2_rules_map.json")
    args = parser.parse_args()

    xlsx = Path(args.input_xlsx)
    assets = Path(args.assets_dir)

    vantagens_payload = json.loads((assets / "vantagens.v3.json").read_text(encoding="utf-8"))
    if isinstance(vantagens_payload, list):
        vantagens = vantagens_payload
    else:
        vantagens = vantagens_payload.get("vantagens", [])
    vantagens_norm = {norm(v.get("nome", "")): v.get("nome", "") for v in vantagens if v.get("nome")}

    ws = load_workbook(xlsx, data_only=True).active

    items = []
    stats = {
        "total": 0,
        "attributeChoice": 0,
        "preReqEmpty": 0,
        "preDefEmpty": 0,
        "modsEmpty": 0,
    }

    for r in range(2, ws.max_row + 1):
        pagina = as_text(ws.cell(r, 1).value)
        nome = as_text(ws.cell(r, 2).value)
        tipo_raw = as_text(ws.cell(r, 3).value)
        predef_raw = as_text(ws.cell(r, 4).value)
        prereq_raw = as_text(ws.cell(r, 5).value)
        desc = as_text(ws.cell(r, 6).value)
        mods = as_text(ws.cell(r, 7).value)

        if not nome:
            continue

        stats["total"] += 1

        tipo = parse_tipo(tipo_raw)
        if tipo["attributeMode"] == "choice":
            stats["attributeChoice"] += 1

        prereq = parse_prerequisitos(prereq_raw, vantagens_norm)
        if prereq["kind"] == "none":
            stats["preReqEmpty"] += 1

        predef = parse_predefinido(predef_raw)
        if predef["kind"] == "empty":
            stats["preDefEmpty"] += 1

        if mods in {"", "-", "—"}:
            stats["modsEmpty"] += 1

        # regra pedida explicitamente pelo usuario para 2 pericias
        nome_n = norm(nome)
        if nome_n in {"pericia profissional", "pericias de passatempo"}:
            tipo["valid"] = True
            tipo["attributeMode"] = "choice"
            tipo["attributeOptions"] = ["DX", "IQ"]
            tipo["reason"] = "regra_usuario_dx_ou_iq"

        item = {
            "row": r,
            "id": slugify(nome),
            "nome": nome,
            "paginaRaw": pagina,
            "tipo": tipo,
            "preRequisito": prereq,
            "preDefinido": predef,
            "descricao": desc,
            "modificadores": {
                "raw": mods,
                "automation": "disabled_if_empty" if mods in {"", "-", "—"} else "manual_text",
            },
            "policy": {
                "emptyPrerequisite": "allow_without_block",
                "emptyPredefinido": "on_zero_points_use_default_predefinido",
                "emptyModifiers": "do_not_automate",
            },
        }

        items.append(item)

    payload = {
        "version": 1,
        "kind": "pericias_v2_rules_map",
        "sourceFile": str(xlsx),
        "summary": stats,
        "notes": [
            "Perícia Profissional (†) e Perícias de Passatempo (†): atributo de escolha do usuário (DX ou IQ).",
            "Pré-requisito vazio: liberar adição sem bloqueio.",
            "Pré-definido vazio: ao adicionar com 0 pontos, usar pré-definido padrão da perícia.",
            "Modificadores vazios: não automatizar nesta etapa (manter texto/manual).",
        ],
        "items": items,
    }

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print("=== Regras estruturadas Perícias V2 ===")
    print(f"input: {xlsx}")
    print(f"output: {out}")
    print(f"total: {stats['total']}")
    print(f"attributeChoice: {stats['attributeChoice']}")
    print(f"preReqEmpty: {stats['preReqEmpty']} | preDefEmpty: {stats['preDefEmpty']} | modsEmpty: {stats['modsEmpty']}")


if __name__ == "__main__":
    main()
