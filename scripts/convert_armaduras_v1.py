#!/usr/bin/env python3
import argparse
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    lowered = ascii_only.lower()
    lowered = re.sub(r"[^a-z0-9]+", "_", lowered)
    lowered = re.sub(r"_+", "_", lowered).strip("_")
    return lowered or "armadura"


def as_text(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_int(s: str):
    if not s:
        return None
    return int(s) if re.fullmatch(r"-?\d+", s) else None


def parse_number_token(token: str, is_money: bool):
    if not token:
        return None

    t = token.strip()
    suffix = t[-1].lower() if t and t[-1].lower() in ("k", "m") else ""
    if suffix:
        t = t[:-1]

    if "," in t and "." in t:
        cleaned = t.replace(".", "").replace(",", ".")
    elif "," in t:
        cleaned = t.replace(",", ".")
    elif "." in t and is_money and re.fullmatch(r"\d{1,3}(?:\.\d{3})+", t):
        cleaned = t.replace(".", "")
    else:
        cleaned = t

    if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", cleaned):
        return None

    value = float(cleaned)
    if suffix == "k":
        value *= 1_000
    elif suffix == "m":
        value *= 1_000_000
    return value


def parse_float_pt(s: str):
    if not s:
        return None
    token_match = re.search(r"[-+]?\d[\d.,]*[kKmM]?", s.strip())
    if not token_match:
        return None
    return parse_number_token(token_match.group(0), is_money=False)


def parse_money(s: str):
    if not s:
        return None
    cleaned = s.replace("$", " ").strip()
    first = re.search(r"[-+]?\d[\d.,]*[kKmM]?", cleaned)
    if not first:
        return None
    return parse_number_token(first.group(0), is_money=True)


def normalize_rd_raw(rd_raw: str) -> str:
    raw = (rd_raw or "").strip()
    if not raw:
        return ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}", raw):
        dt = datetime.fromisoformat(raw)
        return f"{dt.month}/{dt.day}"
    return raw


def is_header_row(col1: str, col2: str):
    return col1.strip().upper() == "NT" and col2.strip().lower() == "armadura"


def is_section_title(col1: str, col2: str):
    return bool(col1) and not col2 and not re.fullmatch(r"-?\d+|\^|â€”", col1.strip())


def convert(input_xlsx: Path, output_json: Path):
    wb = load_workbook(input_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]

    items = []
    used_ids = set()
    current_section = ""
    last_item = None

    for r in range(1, ws.max_row + 1):
        nt_raw = as_text(ws.cell(r, 1).value)
        nome = as_text(ws.cell(r, 2).value)
        local = as_text(ws.cell(r, 3).value)
        rd_raw = normalize_rd_raw(as_text(ws.cell(r, 4).value))
        custo_raw = as_text(ws.cell(r, 5).value)
        peso_raw = as_text(ws.cell(r, 6).value)
        cl_raw = as_text(ws.cell(r, 7).value)
        obs_raw = as_text(ws.cell(r, 8).value)

        if not any([nt_raw, nome, local, rd_raw, custo_raw, peso_raw, cl_raw, obs_raw]):
            continue

        if is_header_row(nt_raw, nome):
            continue
        if is_section_title(nt_raw, nome):
            current_section = nt_raw
            continue

        nt = parse_int(nt_raw)
        is_main_item = bool(nome)

        if is_main_item:
            base_id = slugify(nome)
            final_id = base_id
            suffix = 2
            while final_id in used_ids:
                final_id = f"{base_id}_{suffix}"
                suffix += 1
            used_ids.add(final_id)

            item = {
                "id": final_id,
                "nome": nome,
                "secao": current_section,
                "ntRaw": nt_raw,
                "nt": nt,
                "localRaw": local,
                "rdRaw": rd_raw,
                "custoRaw": custo_raw,
                "custoBase": parse_money(custo_raw),
                "pesoRaw": peso_raw,
                "pesoBaseKg": parse_float_pt(peso_raw),
                "clRaw": cl_raw,
                "cl": parse_int(cl_raw),
                "observacoes": obs_raw,
                "componentes": [],
                "rowNumber": r,
            }
            items.append(item)
            last_item = item
            continue

        if last_item is not None:
            comp = {
                "localRaw": local,
                "rdRaw": rd_raw,
                "custoRaw": custo_raw,
                "custoBase": parse_money(custo_raw),
                "pesoRaw": peso_raw,
                "pesoKg": parse_float_pt(peso_raw),
                "clRaw": cl_raw,
                "cl": parse_int(cl_raw),
                "observacoes": obs_raw,
                "rowNumber": r,
            }
            last_item["componentes"].append(comp)

    payload = {
        "version": 1,
        "kind": "armaduras_v1",
        "sourceFile": str(input_xlsx),
        "generatedAtUtc": datetime.now(timezone.utc).isoformat(),
        "totalItems": len(items),
        "items": items,
    }

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"OK: {len(items)} itens => {output_json}")


def main():
    parser = argparse.ArgumentParser(description="Converte planilha de armaduras para armaduras.v1.json")
    parser.add_argument("--input", required=True, help="Caminho do .xlsx")
    parser.add_argument("--output", required=True, help="Caminho do JSON de saida")
    args = parser.parse_args()
    convert(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
