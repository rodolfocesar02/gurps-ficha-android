#!/usr/bin/env python3
import argparse
import json
import re
import unicodedata
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


EXPECTED = [
    "categoria",
    "nt",
    "nome",
    "tipo_dano",
    "precisao",
    "alcance_distancia",
    "peso",
    "Cdt",
    "tiros",
    "custo",
    "st_minimo",
    "Magnitude",
    "Recuo",
    "Cl",
    "observacoes",
]


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    lowered = ascii_only.lower()
    lowered = re.sub(r"[^a-z0-9]+", "_", lowered)
    lowered = re.sub(r"_+", "_", lowered).strip("_")
    return lowered or "arma_fogo"


def as_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        # OCR em planilha pode converter "x/y" em data. Preservamos "dia/mes".
        return f"{v.day}/{v.month}"
    if isinstance(v, date):
        return f"{v.day}/{v.month}"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def normalize_st(s: str) -> str:
    if not s:
        return ""
    return s.replace("↑", "").replace("â†‘", "").strip()


def convert(input_xlsx: Path, output_json: Path):
    wb = load_workbook(input_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [as_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    headers = [h for h in headers if h]
    if headers[: len(EXPECTED)] != EXPECTED:
        raise SystemExit(f"Cabeçalhos inesperados. Esperado: {EXPECTED} | Obtido: {headers}")

    items = []
    used = set()
    for r in range(2, ws.max_row + 1):
        categoria = as_text(ws.cell(r, 1).value)
        nt = as_text(ws.cell(r, 2).value)
        nome = as_text(ws.cell(r, 3).value)
        dano = as_text(ws.cell(r, 4).value)
        precisao = as_text(ws.cell(r, 5).value)
        alcance = as_text(ws.cell(r, 6).value)
        peso = as_text(ws.cell(r, 7).value)
        cdt = as_text(ws.cell(r, 8).value)
        tiros = as_text(ws.cell(r, 9).value)
        custo = as_text(ws.cell(r, 10).value)
        st = normalize_st(as_text(ws.cell(r, 11).value))
        magnit = as_text(ws.cell(r, 12).value)
        recuo = as_text(ws.cell(r, 13).value)
        cl = as_text(ws.cell(r, 14).value)
        obs = as_text(ws.cell(r, 15).value)

        if not nome:
            continue

        base = slugify(nome)
        fid = base
        n = 2
        while fid in used:
            fid = f"{base}_{n}"
            n += 1
        used.add(fid)

        items.append(
            {
                "id": fid,
                "tipo": "arma_distancia",
                "categoria": "Tabela de Armas de Fogo",
                "grupo": categoria,
                "nt": nt,
                "nome": nome,
                "tipo_dano": dano,
                "precisao": precisao,
                "alcance_distancia": alcance,
                "peso": peso,
                "Cdt": cdt,
                "tiros": tiros,
                "custo": custo,
                "st_minimo": st,
                "Magnitude": magnit,
                "Recuo": recuo,
                "Cl": cl,
                "observacoes": obs,
                "source": "xlsx_armas_fogo",
                "reviewFlags": [],
                "rowNumber": r,
            }
        )

    payload = {
        "version": 1,
        "kind": "armas_fogo_raw_from_xlsx",
        "sourceFile": str(input_xlsx),
        "generatedAtUtc": datetime.now(timezone.utc).isoformat(),
        "totalItems": len(items),
        "items": items,
    }

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"OK: {len(items)} itens => {output_json}")


def main():
    parser = argparse.ArgumentParser(description="Converte tabela de armas de fogo para raw v1")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    convert(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
