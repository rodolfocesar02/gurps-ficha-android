#!/usr/bin/env python3
import argparse
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_HEADERS = [
    "Página",
    "Nome",
    "Dificuldade",
    "Pré-definido",
    "Pré-requisito",
    "Descrição",
    "Modificadores",
]


def as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    lowered = ascii_only.lower()
    lowered = re.sub(r"[^a-z0-9]+", "_", lowered)
    lowered = re.sub(r"_+", "_", lowered).strip("_")
    return lowered or "pericia_artes_marciais"


def parse_page(raw: str):
    if not raw:
        return None
    numbers = re.findall(r"\d+", raw)
    if not numbers:
        return None
    return int(numbers[0])


def convert(input_xlsx: Path, output_json: Path):
    wb = load_workbook(input_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [as_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    headers = headers[: len(EXPECTED_HEADERS)]
    if headers != EXPECTED_HEADERS:
        raise SystemExit(
            "Cabeçalhos inesperados na planilha de perícias de Artes Marciais.\n"
            f"Esperado: {EXPECTED_HEADERS}\n"
            f"Obtido:   {headers}"
        )

    items = []
    used_ids = set()
    for r in range(2, ws.max_row + 1):
        page_raw = as_text(ws.cell(r, 1).value)
        nome = as_text(ws.cell(r, 2).value)
        dificuldade = as_text(ws.cell(r, 3).value)
        pre_definido = as_text(ws.cell(r, 4).value)
        pre_requisito = as_text(ws.cell(r, 5).value)
        descricao = as_text(ws.cell(r, 6).value)
        modificadores = as_text(ws.cell(r, 7).value)

        if not nome:
            continue

        base = slugify(nome)
        item_id = base
        suffix = 2
        while item_id in used_ids:
            item_id = f"{base}_{suffix}"
            suffix += 1
        used_ids.add(item_id)

        items.append(
            {
                "id": item_id,
                "nome": nome,
                "pagina": parse_page(page_raw),
                "paginaRaw": page_raw,
                "dificuldadeRaw": dificuldade,
                "preDefinidoRaw": pre_definido,
                "preRequisitoRaw": pre_requisito,
                "descricao": descricao,
                "modificadores": modificadores,
                "sourceBook": "Artes Marciais",
                "sourceFile": input_xlsx.name,
            }
        )

    payload = {
        "version": 1,
        "kind": "pericias_artes_marciais_v1",
        "sourceFile": str(input_xlsx),
        "generatedAtUtc": datetime.now(timezone.utc).isoformat(),
        "totalItems": len(items),
        "items": items,
    }

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"OK: {len(items)} itens => {output_json}")


def main():
    parser = argparse.ArgumentParser(description="Converte perícias de Artes Marciais para JSON v1")
    parser.add_argument("--input", required=True, help="Caminho do XLSX")
    parser.add_argument("--output", required=True, help="Caminho do JSON de saída")
    args = parser.parse_args()
    convert(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
