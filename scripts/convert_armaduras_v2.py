#!/usr/bin/env python3
import argparse
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

MOJIBAKE_MARKERS = ("Ã", "Â", "�")


LOW_NOTES = {
    1: "Pode ser ocultado como ou sob uma peça de roupa.",
    2: "Cobertura parcial: sandálias concedem RD 1 para as solas dos pés, enquanto as placas de peito protegem apenas contra ataques frontais.",
    3: "RD dividida: utilize a menor RD contra ataques por contusão.",
    4: "O elmo impõe a desvantagem Sem Visão Periférica (pág. 160).",
}

HIGH_GLOBAL = (
    "Todos os aparelhos eletrônicos e sistemas elétricos (incluindo os trajes de combate) das armaduras de NT7+ "
    "funcionam durante (NT - 6) × 6 horas, antes de precisarem ser recarregados ou reabastecidos."
)

HIGH_NOTES = {
    1: "RD dividida: use a primeira RD, mais alta, contra ataques perfurantes e de corte; utilize a segunda RD, mais baixa, contra todos os outros tipos de dano.",
    2: "Pode ser considerada uma roupa ou ocultada com ou sob as roupas.",
    3: "A RD aumenta junto com o NT. Após o NT inicial: Inicial×1; Inicial+1×1,5; Inicial+2×2; Inicial+3×3; superior×4.",
    4: "Sensores biomédicos permitem monitoramento remoto dos sinais vitais, concedendo +1 em Diagnose ao examinar o usuário. O traje é climatizado.",
    5: "Exige a perícia Traje NBQ; em NT9+, não limita DX. Com máscara/capacete [7], concede a vantagem Lacrado.",
    6: "RD dividida: use a RD mais alta somente se o ataque atingir tronco (blindagem pessoal), crânio (capacete) ou sola do pé (calçado).",
    7: "Inclui Pulmões com Filtro, Olfato Protegido e Visão Protegida; antes do NT9 também impõe Sem Visão Periférica.",
    8: "Concede Visão Protegida.",
    9: "A RD do traje funciona apenas contra dano por queimadura ou corrosão.",
    10: "Exige perícia Traje Pressurizado. Com capacete, concede Não Respira (12h), Olfato Protegido, Lacrado e Resistência ao Vácuo.",
    11: "Concede Audição Protegida, Visão Protegida e Rádio. Em NT9+, concede Senso de Direção (Requer Sinal), Infravisão, Visão Noturna 9 e (NT-8) níveis de Visão Telescópica.",
    12: "Exige perícia Traje de Combate. Concede ST de Levantamento/ST de Golpe +10 e Super Salto 1; melhora com NT. Com elmo, concede Não Respira (12h), Olfato Protegido, Lacrado e Resistência ao Vácuo.",
}


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    lowered = ascii_only.lower()
    lowered = re.sub(r"[^a-z0-9]+", "_", lowered)
    lowered = re.sub(r"_+", "_", lowered).strip("_")
    return lowered or "armadura"


def as_text(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return repair_mojibake(str(value).strip())


def repair_mojibake(text: str) -> str:
    current = text
    for _ in range(2):
        if not any(marker in current for marker in MOJIBAKE_MARKERS):
            break
        try:
            repaired = current.encode("latin-1").decode("utf-8")
        except UnicodeError:
            break
        if repaired == current:
            break
        current = repaired
    return current


def parse_int(s: str):
    if not s:
        return None
    return int(s) if re.fullmatch(r"-?\d+", s) else None


def parse_number_token(token: str, is_money: bool):
    if not token:
        return None
    t = token.strip()
    is_plus = t.startswith("+")
    if is_plus:
        t = t[1:].strip()
    suffix = t[-1].lower() if t and t[-1].lower() in ("k", "m") else ""
    if suffix:
        t = t[:-1]
    if "," in t and "." in t:
        cleaned = t.replace(".", "").replace(",", ".")
    elif "," in t:
        cleaned = t.replace(",", ".")
    elif "." in t and is_money and re.fullmatch(r"\d{1,3}(?:\.\d{3})+", t):
        cleaned = t.replace(".", "")
    else:
        cleaned = t
    if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?", cleaned):
        return None
    value = float(cleaned)
    if suffix == "k":
        value *= 1_000
    elif suffix == "m":
        value *= 1_000_000
    return value


def parse_float_pt(s: str):
    if not s:
        return None
    token_match = re.search(r"[-+]?\d[\d.,]*[kKmM]?", s.strip())
    if not token_match:
        return None
    return parse_number_token(token_match.group(0), is_money=False)


def parse_money(s: str):
    if not s:
        return None
    cleaned = s.replace("$", " ").strip()
    first = re.search(r"[-+]?\d[\d.,]*[kKmM]?", cleaned)
    if not first:
        return None
    return parse_number_token(first.group(0), is_money=True)


def extract_note_refs(obs_raw: str):
    if not obs_raw:
        return []
    refs = re.findall(r"\[(\d+)\]", obs_raw)
    return sorted({int(x) for x in refs})


def normalize_local_token(s: str) -> str:
    n = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii").lower().strip()
    n = re.sub(r"\s+", "_", n)
    return n


def split_locais(raw: str):
    return [x.strip() for x in re.split(r"[,;/|]", raw or "") if x.strip()]


def expand_macro_local(local_token: str):
    t = normalize_local_token(local_token)
    if t == "cabeca":
        return ["cranio", "rosto", "olhos"]
    if t == "corpo":
        return ["pescoco", "tronco", "virilha"]
    if t == "membros":
        return ["bracos", "pernas"]
    if t == "traje_completo":
        return ["pescoco", "tronco", "virilha", "bracos", "maos", "pernas", "pes"]
    return [t]


def parse_rd(raw: str):
    txt = (raw or "").strip()
    flexivel = "*" in txt
    frontal = "D" in txt.upper()
    cleaned = txt.replace("*", "").replace("D", "").replace("d", "").strip()
    if "/" in cleaned:
        a, b = cleaned.split("/", 1)
        return {
            "raw": txt,
            "principal": parse_int(a.strip()),
            "secundario": parse_int(b.strip()),
            "flexivel": flexivel,
            "frontalSomente": frontal,
            "dividida": True,
        }
    return {
        "raw": txt,
        "principal": parse_int(cleaned),
        "secundario": None,
        "flexivel": flexivel,
        "frontalSomente": frontal,
        "dividida": False,
    }


def note_texts(note_refs, nt):
    if nt is not None and nt >= 6:
        note_map = HIGH_NOTES
        texts = [repair_mojibake(HIGH_GLOBAL)] if note_refs else []
    else:
        note_map = LOW_NOTES
        texts = []
    for ref in note_refs:
        txt = note_map.get(ref)
        if txt:
            texts.append(repair_mojibake(txt))
    return texts


def build_tags(local_raw: str, rd_meta: dict, nt: int | None, refs: list[int], nome: str):
    tags = set()
    for l in split_locais(local_raw):
        tags.add(f"local:{normalize_local_token(l)}")
        for ex in expand_macro_local(l):
            tags.add(f"local_exp:{ex}")
    if rd_meta.get("flexivel"):
        tags.add("rd:flexivel")
    if rd_meta.get("frontalSomente"):
        tags.add("rd:frontal")
    if rd_meta.get("dividida"):
        tags.add("rd:dividida")
    if nt is not None:
        tags.add(f"nt:{nt}")
        tags.add("nt:alto" if nt >= 6 else "nt:baixo")
    for ref in refs:
        tags.add(f"obs:{ref}")
    if nome.strip().startswith("+"):
        tags.add("tipo:addon")
    else:
        tags.add("tipo:base")
    return sorted(tags)


def convert(input_xlsx: Path, output_json: Path):
    wb = load_workbook(input_xlsx, data_only=True)
    items = []
    used_ids = set()

    for ws in wb.worksheets:
        last_item = None
        for r in range(2, ws.max_row + 1):
            nt_raw = as_text(ws.cell(r, 1).value)
            nome = as_text(ws.cell(r, 2).value)
            local = as_text(ws.cell(r, 3).value)
            rd_raw = as_text(ws.cell(r, 4).value)
            custo_raw = as_text(ws.cell(r, 5).value)
            peso_raw = as_text(ws.cell(r, 6).value)
            cl_raw = as_text(ws.cell(r, 7).value)
            obs_raw = as_text(ws.cell(r, 8).value)

            if not any([nt_raw, nome, local, rd_raw, custo_raw, peso_raw, cl_raw, obs_raw]):
                continue

            nt = parse_int(nt_raw)
            is_addon = nome.startswith("+")
            refs = extract_note_refs(obs_raw)
            rd_meta = parse_rd(rd_raw)

            if nome and not is_addon:
                base_id = slugify(nome)
                final_id = base_id
                suffix = 2
                while final_id in used_ids:
                    final_id = f"{base_id}_{suffix}"
                    suffix += 1
                used_ids.add(final_id)

                entry = {
                    "id": final_id,
                    "nome": nome,
                    "secao": ws.title,
                    "ntRaw": nt_raw,
                    "nt": nt,
                    "localRaw": local,
                    "locaisNorm": sorted({x for t in split_locais(local) for x in expand_macro_local(t)}),
                    "rdRaw": rd_raw,
                    "rd": rd_meta,
                    "custoRaw": custo_raw,
                    "custoBase": parse_money(custo_raw),
                    "pesoRaw": peso_raw,
                    "pesoBaseKg": parse_float_pt(peso_raw),
                    "clRaw": cl_raw,
                    "cl": parse_int(cl_raw),
                    "observacoesRaw": obs_raw,
                    "observacoesRefs": refs,
                    "observacoesDetalhadas": note_texts(refs, nt),
                    "tags": build_tags(local, rd_meta, nt, refs, nome),
                    "componentes": [],
                    "rowNumber": r,
                }
                items.append(entry)
                last_item = entry
                continue

            if is_addon and last_item is not None:
                comp = {
                    "nomeRaw": nome,
                    "localRaw": local,
                    "locaisNorm": sorted({x for t in split_locais(local) for x in expand_macro_local(t)}),
                    "rdRaw": rd_raw,
                    "rd": rd_meta,
                    "custoRaw": custo_raw,
                    "custoBase": parse_money(custo_raw),
                    "pesoRaw": peso_raw,
                    "pesoKg": parse_float_pt(peso_raw),
                    "clRaw": cl_raw,
                    "cl": parse_int(cl_raw),
                    "observacoesRaw": obs_raw,
                    "observacoesRefs": refs,
                    "observacoesDetalhadas": note_texts(refs, last_item.get("nt")),
                    "tags": build_tags(local, rd_meta, last_item.get("nt"), refs, nome),
                    "rowNumber": r,
                }
                last_item["componentes"].append(comp)

    payload = {
        "version": 2,
        "kind": "armaduras_v2",
        "sourceFile": str(input_xlsx),
        "generatedAtUtc": datetime.now(timezone.utc).isoformat(),
        "totalItems": len(items),
        "notes": {
            "baixoNt": {str(k): v for k, v in LOW_NOTES.items()},
            "altoNtGlobal": HIGH_GLOBAL,
            "altoNt": {str(k): v for k, v in HIGH_NOTES.items()},
        },
        "items": items,
    }

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"OK: {len(items)} itens => {output_json}")


def main():
    parser = argparse.ArgumentParser(description="Converte planilha de armaduras para armaduras.v2.json")
    parser.add_argument("--input", required=True, help="Caminho do .xlsx")
    parser.add_argument("--output", required=True, help="Caminho do JSON de saida")
    args = parser.parse_args()
    convert(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
