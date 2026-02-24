#!/usr/bin/env python3
import argparse
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


ALLOWED_COST_KINDS = {"fixed", "perLevel", "choice", "range", "special"}


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    lowered = ascii_only.lower()
    lowered = re.sub(r"[^a-z0-9]+", "_", lowered)
    lowered = re.sub(r"_+", "_", lowered).strip("_")
    return lowered or "desvantagem"


def to_int(value):
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        cleaned = value.replace("*", "").strip()
        m = re.fullmatch(r"\s*(-?\d+)\s*", cleaned)
        if m:
            return int(m.group(1))
    return None


def ensure_negative(n: int) -> int:
    return n if n <= 0 else -n


def parse_cost(raw_cost, name_slug: str):
    text = "" if raw_cost is None else str(raw_cost).strip()
    lower = text.lower().replace("–", "-").replace("*", "")
    lower = re.sub(r"\s+", " ", lower).strip()

    result = {
        "costKind": "special",
        "fixed": None,
        "perLevel": None,
        "options": None,
        "min": None,
        "max": None,
        "rawCost": text,
        "specialRule": name_slug,
    }

    fixed = to_int(raw_cost)
    if fixed is not None:
        result.update(
            {
                "costKind": "fixed",
                "fixed": ensure_negative(fixed),
                "specialRule": None,
            }
        )
        return result

    m = re.fullmatch(r"(-?\d+)\s*/\s*n[ií]vel", lower)
    if m:
        result.update(
            {
                "costKind": "perLevel",
                "perLevel": ensure_negative(int(m.group(1))),
                "specialRule": None,
            }
        )
        return result

    m = re.fullmatch(r"(-?\d+)\s*a\s*(-?\d+)", lower)
    if m:
        a = ensure_negative(int(m.group(1)))
        b = ensure_negative(int(m.group(2)))
        lo, hi = sorted([a, b])
        result.update(
            {
                "costKind": "range",
                "min": lo,
                "max": hi,
                "specialRule": None,
            }
        )
        return result

    # "x ou y" / "x, y ou z"
    if " ou " in lower:
        nums = [ensure_negative(int(x)) for x in re.findall(r"-?\d+", lower)]
        if len(nums) >= 2:
            result.update(
                {
                    "costKind": "choice",
                    "options": sorted(set(nums)),
                    "specialRule": None,
                }
            )
            return result

    # Casos especiais com custo por "unidade" (ex.: "-5/contato")
    m = re.fullmatch(r"(-?\d+)\s*/\s*[a-zà-ÿ_]+", lower)
    if m:
        result.update(
            {
                "costKind": "special",
                "fixed": ensure_negative(int(m.group(1))),
            }
        )
        return result

    if "vari" in lower:
        return result

    nums = [ensure_negative(int(x)) for x in re.findall(r"-?\d+", lower)]
    if len(nums) == 1:
        result.update({"costKind": "fixed", "fixed": nums[0], "specialRule": None})
    elif len(nums) >= 2:
        # fallback seguro: múltiplos valores numéricos => escolha
        result.update(
            {
                "costKind": "choice",
                "options": sorted(set(nums)),
                "specialRule": None,
            }
        )
    return result


def guess_tags(name: str):
    n = name.lower()
    tags = set()

    if any(k in n for k in ("mana", "mág", "magia", "arcano", "arcana", "feiti", "encanta")):
        tags.add("magica")
    if any(k in n for k in ("social", "status", "reputa", "odioso", "intoler", "inimigo", "dependente")):
        tags.add("social")
    if any(k in n for k in ("fobia", "compuls", "obsess", "delirio", "loucura", "mente", "mental")):
        tags.add("mental")
    if any(k in n for k in ("cego", "surdo", "mudo", "doenca", "aleij", "fisic", "dor", "fraco")):
        tags.add("fisica")
    if any(k in n for k in ("combate", "sangue", "ferimento", "vulner", "pacif")):
        tags.add("combate")

    return sorted(tags)


def validate_items(items):
    errors = []
    seen = set()
    for i, o in enumerate(items):
        for field in (
            "id",
            "nome",
            "pagina",
            "costKind",
            "fixed",
            "perLevel",
            "options",
            "min",
            "max",
            "rawCost",
            "specialRule",
            "tags",
        ):
            if field not in o:
                errors.append(f"[{i}] missing field: {field}")
        if o["id"] in seen:
            errors.append(f"[{i}] duplicate id: {o['id']}")
        seen.add(o["id"])
        if o["costKind"] not in ALLOWED_COST_KINDS:
            errors.append(f"[{i}] invalid costKind: {o['costKind']}")
        if o["costKind"] == "fixed" and o["fixed"] is None:
            errors.append(f"[{i}] fixed sem fixed: {o['id']}")
        if o["costKind"] == "perLevel" and o["perLevel"] is None:
            errors.append(f"[{i}] perLevel sem perLevel: {o['id']}")
    return errors


def convert(input_xlsx: Path, output_json: Path):
    wb = load_workbook(input_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]

    items = []
    used_ids = set()

    for r in range(2, ws.max_row + 1):
        page = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        cost = ws.cell(r, 3).value

        if not name:
            continue

        name = str(name).strip()
        if not name:
            continue

        if isinstance(page, (int, float)):
            page = int(page)
        else:
            page = to_int(str(page)) or 0

        base_id = slugify(name)
        final_id = base_id
        suffix = 2
        while final_id in used_ids:
            final_id = f"{base_id}_{suffix}"
            suffix += 1
        used_ids.add(final_id)

        parsed = parse_cost(cost, final_id)

        item = {
            "id": final_id,
            "nome": name,
            "pagina": page,
            "costKind": parsed["costKind"],
            "fixed": parsed["fixed"],
            "perLevel": parsed["perLevel"],
            "options": parsed["options"],
            "min": parsed["min"],
            "max": parsed["max"],
            "rawCost": parsed["rawCost"],
            "specialRule": parsed["specialRule"],
            "tags": guess_tags(name),
        }
        items.append(item)

    errors = validate_items(items)
    if errors:
        raise SystemExit("Validation errors:\n" + "\n".join(errors[:200]))

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(items, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    stats = {}
    for it in items:
        stats[it["costKind"]] = stats.get(it["costKind"], 0) + 1
    print(f"OK: {len(items)} itens => {output_json}")
    print(f"Distribuição costKind: {stats}")


def main():
    parser = argparse.ArgumentParser(description="Converte planilha de desvantagens para desvantagens.v2.json")
    parser.add_argument("--input", required=True, help="Caminho do .xlsx")
    parser.add_argument("--output", required=True, help="Caminho do desvantagens.v2.json")
    args = parser.parse_args()
    convert(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
