#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from pypdf import PdfReader


def normalize(text: str) -> str:
    txt = unicodedata.normalize("NFKD", str(text))
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    txt = txt.lower()
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def normalize_key(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", normalize(text)).strip()


@dataclass
class TecnicaXlsx:
    nome: str
    pagina: Optional[int]
    descricao: str


class XlsxIndex:
    def __init__(self) -> None:
        self.by_name_page: Dict[Tuple[str, Optional[int]], List[TecnicaXlsx]] = {}
        self.by_name: Dict[str, List[TecnicaXlsx]] = {}

    def add(self, item: TecnicaXlsx) -> None:
        k_name = normalize_key(item.nome)
        self.by_name_page.setdefault((k_name, item.pagina), []).append(item)
        self.by_name.setdefault(k_name, []).append(item)

    def find(self, nome: str, pagina: Optional[int]) -> Optional[TecnicaXlsx]:
        k_name = normalize_key(nome)
        exact = self.by_name_page.get((k_name, pagina), [])
        if len(exact) == 1:
            return exact[0]
        if len(exact) > 1:
            descs = {x.descricao for x in exact}
            if len(descs) == 1:
                return exact[0]
        by_name = self.by_name.get(k_name, [])
        if len(by_name) == 1:
            return by_name[0]
        if len(by_name) > 1 and pagina is not None:
            same_page = [x for x in by_name if x.pagina == pagina]
            if len(same_page) == 1:
                return same_page[0]
            if len(same_page) > 1:
                descs = {x.descricao for x in same_page}
                if len(descs) == 1:
                    return same_page[0]
        return None


def header_map(headers: List[str]) -> Dict[str, int]:
    normalized = {normalize_key(h): i for i, h in enumerate(headers)}

    def pick(*aliases: str) -> int:
        for a in aliases:
            if a in normalized:
                return normalized[a]
        raise KeyError(f"Coluna ausente: {aliases}")

    return {
        "pagina": pick("pagina"),
        "nome": pick("nome"),
        "descricao": pick("descricao"),
    }


def parse_int(raw) -> Optional[int]:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.lower() == "nan":
        return None
    m = re.search(r"\d+", s)
    return int(m.group(0)) if m else None


def load_xlsx(path: Path) -> XlsxIndex:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = ["" if ws.cell(1, c).value is None else str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    idx = header_map(headers)
    out = XlsxIndex()

    for r in range(2, ws.max_row + 1):
        nome = ws.cell(r, idx["nome"] + 1).value
        if nome is None:
            continue
        nome_s = str(nome).strip()
        if not nome_s:
            continue
        desc = ws.cell(r, idx["descricao"] + 1).value
        if desc is None:
            continue
        desc_s = str(desc).replace("\xa0", " ").strip()
        if not desc_s:
            continue
        page = parse_int(ws.cell(r, idx["pagina"] + 1).value)
        out.add(TecnicaXlsx(nome=nome_s, pagina=page, descricao=desc_s))

    return out


def load_pdf_text(pdf_path: Path) -> str:
    reader = PdfReader(str(pdf_path))
    chunks: List[str] = []
    for page in reader.pages:
        txt = page.extract_text() or ""
        if txt:
            chunks.append(txt)
    return normalize("\n".join(chunks))


def find_snippet(text: str, name_key: str, size: int = 3500) -> str:
    idx = text.find(name_key)
    if idx < 0:
        return ""
    start = max(0, idx - 300)
    end = min(len(text), idx + size)
    return text[start:end]


def description_similarity(desc: str, snippet: str) -> float:
    if not desc or not snippet:
        return 0.0
    d = normalize(desc)
    if len(d) > 1800:
        d = d[:1800]
    return SequenceMatcher(None, d, snippet).ratio()


def main() -> None:
    parser = argparse.ArgumentParser(description="Merge das descricoes de tecnicas via xlsx + auditoria com pdf")
    parser.add_argument("--json", required=True, help="JSON de tecnicas (tecnicas.v1.json)")
    parser.add_argument("--xlsx-artes", required=True)
    parser.add_argument("--xlsx-gunfu", required=True)
    parser.add_argument("--pdf-artes", required=True)
    parser.add_argument("--pdf-gunfu", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--inplace", action="store_true")
    args = parser.parse_args()

    json_path = Path(args.json)
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    items = payload.get("items", [])

    idx_artes = load_xlsx(Path(args.xlsx_artes))
    idx_gunfu = load_xlsx(Path(args.xlsx_gunfu))

    pdf_artes = load_pdf_text(Path(args.pdf_artes))
    pdf_gunfu = load_pdf_text(Path(args.pdf_gunfu))

    updated = 0
    unchanged = 0
    no_match: List[dict] = []
    pdf_name_missing: List[dict] = []
    possible_divergence: List[dict] = []

    for item in items:
        src = str(item.get("sourceBook", "")).strip()
        nome = str(item.get("nome", "")).strip()
        pagina = item.get("pagina") if isinstance(item.get("pagina"), int) else None

        idx = idx_artes if src == "Artes Marciais" else idx_gunfu if src == "Gun Fu" else None
        if idx is None:
            continue

        found = idx.find(nome, pagina)
        if not found:
            no_match.append({"id": item.get("id"), "nome": nome, "pagina": pagina, "sourceBook": src})
            continue

        before = str(item.get("descricao", ""))
        after = found.descricao
        if normalize(before) != normalize(after):
            item["descricao"] = after
            updated += 1
        else:
            unchanged += 1

        name_key = normalize_key(nome)
        text = pdf_artes if src == "Artes Marciais" else pdf_gunfu
        snippet = find_snippet(text, name_key)
        if not snippet:
            pdf_name_missing.append({"id": item.get("id"), "nome": nome, "sourceBook": src})
            continue

        score = description_similarity(after, snippet)
        if score < 0.22:
            possible_divergence.append(
                {
                    "id": item.get("id"),
                    "nome": nome,
                    "sourceBook": src,
                    "score": round(score, 4),
                }
            )

    payload["totalItems"] = len(items)

    if args.inplace:
        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    report = {
        "json": str(json_path),
        "total_items": len(items),
        "updated_descricao": updated,
        "unchanged_descricao": unchanged,
        "no_match_count": len(no_match),
        "pdf_name_missing_count": len(pdf_name_missing),
        "possible_divergence_count": len(possible_divergence),
        "no_match": no_match,
        "pdf_name_missing": pdf_name_missing,
        "possible_divergence": possible_divergence,
    }
    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
